# -*- coding: utf-8 -*-

# -------------------------------------------------------------------------------
# Name:         test1
# Description:  遍历Excel指定行并打印单元格列定位和公式
# Author:       shaver
# Date:         2025/9/9
# -------------------------------------------------------------------------------
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter  # 导入获取列字母的工具函数

excel_path = '/Users/shaver/Desktop/data/模板.xlsx'
# 子表
# 业务报告
# WT报告
# SP
# SD
# SB
# Base
sheet_name = 'Base'

# 指定要遍历的行号（注意：openpyxl中的行号从1开始计数）
target_row = 1  # 你可以修改这个值来指定不同的行

try:
    # 加载 Excel 工作簿，设置 data_only=False 以获取公式而非计算值
    wb = load_workbook(excel_path, data_only=False)
    ws = wb[sheet_name]  # 选择工作表

    print(f"开始遍历工作表 '{sheet_name}' 的第 {target_row} 行:")

    # 遍历指定行的所有单元格
    for col_index in range(1, ws.max_column + 1):  # 从第1列到最大列
        cell = ws.cell(row=target_row, column=col_index)
        column_letter = get_column_letter(col_index)  # 获取列的字母标识，如A, B, C[7](@ref)

        # 检查单元格是否包含公式
        if cell.data_type == 'f':  # 'f' 表示单元格包含公式[6](@ref)
            # 如果value是对象取text属性，否则直接取value
            formula = cell.value.text if hasattr(cell.value, 'text') else cell.value
            print(f"单元格 {column_letter}{target_row}: 公式 = {formula}")
        else:
            # 值去掉换行符
            value = str(cell.value).replace('\n', '')
            print(f"{column_letter}{target_row},{value}")

    print("遍历完成。")

except FileNotFoundError:
    print(f"错误：找不到文件 '{excel_path}'，请检查文件路径是否正确")
except KeyError:
    print(f"错误：工作簿中不存在名为 '{sheet_name}' 的工作表")
except Exception as e:
    print(f"处理过程中发生错误：{str(e)}")