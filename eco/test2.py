# -*- coding: utf-8 -*-

# -------------------------------------------------------------------------------
# Name:         test1
# Description:  遍历Excel指定行并写入单元格列定位和公式到新的Excel文件
# Author:       shaver
# Date:         2025/9/9
# -------------------------------------------------------------------------------
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

excel_path = '/Users/shaver/Desktop/data/模板.xlsx'
# 可用的sheet名称:
# 子表
# 业务报告
# WT报告
# SP
# SD
# SB
# Base
sheet_name = '子表'

# 指定要遍历的行号（注意：openpyxl中的行号从1开始计数）
target_row = 5  # 你可以修改这个值来指定不同的行

# 输出Excel文件路径
output_excel_path = 'out1.xlsx'

try:
    # 加载 Excel 工作簿，设置 data_only=False 以获取公式而非计算值
    wb = load_workbook(excel_path, data_only=False)
    ws = wb[sheet_name]  # 选择工作表

    # 创建新的工作簿用于保存结果[4,8](@ref)
    result_wb = Workbook()
    result_ws = result_wb.active
    result_ws.title = f"{sheet_name}_第{target_row}行结果"

    # 设置表头
    result_ws['A1'] = '单元格位置'
    result_ws['B1'] = '类型'
    result_ws['C1'] = '内容'
    result_ws['D1'] = '列索引'
    result_ws['E1'] = '行索引'

    # 设置表头样式[4](@ref)
    from openpyxl.styles import Font

    bold_font = Font(bold=True)
    for cell in result_ws[1]:
        cell.font = bold_font

    print(f"开始遍历工作表 '{sheet_name}' 的第 {target_row} 行:")

    # 从第二行开始写入数据
    row_count = 2

    # 遍历指定行的所有单元格
    for col_index in range(1, ws.max_column + 1):  # 从第1列到最大列
        cell = ws.cell(row=target_row, column=col_index)
        column_letter = get_column_letter(col_index)  # 获取列的字母标识，如A, B, C

        # 检查单元格是否包含公式
        if cell.data_type == 'f':  # 'f' 表示单元格包含公式
            # 如果是对象获取属性text 否则获取值
            formula = cell.value.text if hasattr(cell.value, 'text') else cell.value
            # 去除公式前面的等号
            if formula.startswith('='):
                formula = formula[1:]
            cell_position = f"{column_letter}{target_row}"

            # 在控制台输出
            print(f"单元格 {cell_position}: 公式 = {formula}")

            # 写入到结果Excel文件[6,8](@ref)
            result_ws.cell(row=row_count, column=1, value=cell_position)
            result_ws.cell(row=row_count, column=2, value="公式")
            result_ws.cell(row=row_count, column=3, value=str(formula))
            result_ws.cell(row=row_count, column=4, value=col_index)
            result_ws.cell(row=row_count, column=5, value=target_row)

        else:
            value = cell.value
            cell_position = f"{column_letter}{target_row}"

            # 在控制台输出
            print(f"单元格 {cell_position}: 值 = {value}")

            # 写入到结果Excel文件
            result_ws.cell(row=row_count, column=1, value=cell_position)
            result_ws.cell(row=row_count, column=2, value="值")
            result_ws.cell(row=row_count, column=3, value=str(value) if value is not None else "空值")
            result_ws.cell(row=row_count, column=4, value=col_index)
            result_ws.cell(row=row_count, column=5, value=target_row)

        row_count += 1

    # 自动调整列宽[4](@ref)
    for column in result_ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        result_ws.column_dimensions[column_letter].width = adjusted_width

    # 保存结果文件[8](@ref)
    result_wb.save(output_excel_path)
    print(f"\n遍历完成。结果已保存到Excel文件: {output_excel_path}")

except FileNotFoundError:
    print(f"错误：找不到文件 '{excel_path}'，请检查文件路径是否正确")
except KeyError:
    print(f"错误：工作簿中不存在名为 '{sheet_name}' 的工作表")
except Exception as e:
    print(f"处理过程中发生错误：{str(e)}")